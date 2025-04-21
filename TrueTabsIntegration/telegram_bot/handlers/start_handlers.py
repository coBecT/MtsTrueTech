import asyncio
import logging
import sys
from typing import Union
from aiogram import Router, F
from aiogram.filters import CommandStart, StateFilter
from aiogram.fsm.context import FSMContext # Импортируем для очистки состояния при отмене
import json
import io
import csv
import httpx
from aiogram.types import CallbackQuery, FSInputFile, Message, InlineKeyboardMarkup, InlineKeyboardButton
from telegram_bot.keyboards.inline_with_export_update import main_menu_keyboard
from telegram_bot.database import get_tt_config
from telegram_bot.utils.encryption import decrypt_data


from ..keyboards.inline import main_menu_keyboard # Импортируем клавиатуру главного меню
# from .upload_handlers import UploadProcess # Импортируем классы состояний для очистки при отмене
# from .config_handlers import ConfigProcess
# from .scheduled_handlers import ScheduleProcess # Импортируем классы состояний для очистки при отмене

logger = logging.getLogger(__name__) # Логгер для start_handlers

router = Router()

# Хэндлер на команду /start
@router.message(CommandStart())
async def command_start_handler(message: Message, state: FSMContext) -> None:
    """
    Обрабатывает команду /start. Приветствует пользователя и показывает главное меню.
    Также очищает текущее состояние FSM, если оно есть.
    """
    await state.clear() # Очищаем любое текущее состояние FSM при старте

    welcome_text = (
        f"Привет, {message.from_user.full_name}! 👋\n"
        f"Я бот для автоматизации загрузки данных в True Tabs из различных источников с помощью Rust утилиты.\n\n"
        f"Выберите действие:"
    )
    await message.answer(welcome_text, reply_markup=main_menu_keyboard())
    logger.info(f"Received /start command from user {message.from_user.id}")

# Общий хэндлер для кнопки "❌ Отмена" или callback "cancel"
# Он должен быть определен на уровне диспетчера или в роутере, который включен в диспетчер
# и срабатывает для всех состояний или без состояний, кроме специфических.
# В данном случае, добавим его в start_handlers, т.к. он связан с начальным меню и отменой.
# Removed the handler for callback "main_menu" to delete the menu as requested
@router.message(F.text.lower() == "отмена", StateFilter("*")) # Сработает на текст "отмена" в любом состоянии
async def cancel_fsm_process(callback_or_message: Union[Message, CallbackQuery], state: FSMContext):
    """
    Обрабатывает команду или callback для отмены текущего FSM процесса.
    Возвращает пользователя в главное меню.
    """
    current_state = await state.get_state()
    if current_state is None:
        # Если состояния нет, просто показываем главное меню (для текстовой команды)
        if isinstance(callback_or_message, Message):
             await callback_or_message.answer("Нет активной операции для отмены.", reply_markup=main_menu_keyboard())
        else: # Если это CallbackQuery без состояния, просто отвечаем и показываем меню
             await callback_or_message.message.edit_text("Нет активной операции для отмены.", reply_markup=main_menu_keyboard())
             await callback_or_message.answer() # Отвечаем на callback

        return # Выходим, если состояния не было

    logger.info(f"Cancelling FSM process. User ID: {callback_or_message.from_user.id}, State: {current_state}")

    await state.clear() # Очищаем все данные и состояние FSM

    message_text = "Операция отменена."
    reply_markup = main_menu_keyboard()

    if isinstance(callback_or_message, CallbackQuery):
        try:
            # Пытаемся отредактировать сообщение callback'а, если это возможно
            await callback_or_message.message.edit_text(message_text, reply_markup=reply_markup)
        except TelegramAPIError:
             # Если сообщение слишком старое для редактирования, отправляем новое
             await callback_or_message.message.answer(message_text, reply_markup=reply_markup)

        await callback_or_message.answer("Отменено") # Отвечаем на callback

    else: # Это текстовое сообщение "отмена"
        await callback_or_message.answer(message_text, reply_markup=reply_markup)

@router.message(StateFilter("*"))
async def handle_unexpected_message(message: Message, state: FSMContext):
    current_state = await state.get_state()
    if current_state:
        await message.answer(f"Бот ожидает другой ввод в текущем состоянии ({current_state}).\n"
                             "Для отмены текущей операции нажмите '❌ Отмена'.")
    else:
        # Если состояния нет, это просто обычное сообщение, которое бот не понимает
        await message.answer("Извините, я не понял вашу команду. Используйте меню.")

@router.callback_query(F.data == "export_data")
async def export_data_handler(callback: CallbackQuery, state: FSMContext):
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="Excel (.xlsx)", callback_data="export_format:xlsx"),
            InlineKeyboardButton(text="CSV (.csv)", callback_data="export_format:csv"),
        ],
        [
            InlineKeyboardButton(text="❌ Отмена", callback_data="cancel")
        ]
    ])
    await callback.message.edit_text("Выберите формат для выгрузки данных:", reply_markup=keyboard)
    await state.set_state("export_select_format")

@router.callback_query(F.data.startswith("export_format:"))
async def export_format_handler(callback: CallbackQuery, state: FSMContext):
    format_selected = callback.data.split(":")[1]
    data = await fetch_truetabs_data(callback.from_user.id)
    if data is None:
        await callback.message.answer("Ошибка при получении данных из True Tabs.")
        await callback.answer()
        return

    if format_selected == "csv":
        file_bytes = convert_to_csv_bytes(data)
        file_name = "exported_data.csv"
    else:
        file_bytes = convert_to_excel_bytes(data)
        file_name = "exported_data.xlsx"

    file = FSInputFile(io.BytesIO(file_bytes), filename=file_name)
    await callback.message.answer_document(file)
    await callback.message.edit_text("Выгрузка завершена.", reply_markup=main_menu_keyboard())
    await state.clear()
    await callback.answer()

@router.callback_query(F.data == "update_data")
async def update_data_handler(callback: CallbackQuery):
    success = await send_update_to_truetabs(callback.from_user.id)
    if success:
        await callback.message.answer("Данные успешно обновлены через API.")
    else:
        await callback.message.answer("Ошибка при обновлении данных через API.")
    await callback.message.edit_text("Основное меню:", reply_markup=main_menu_keyboard())
    await callback.answer()

async def fetch_truetabs_data(user_id: int):
    tt_config = await get_tt_config(user_id)
    if not tt_config:
        return None
    api_token = decrypt_data(tt_config['api_token'])
    datasheet_id = tt_config['datasheet_id']

    url = f"https://true.tabs.sale/fusion/v1/datasheets/{datasheet_id}/records?viewId=viwyshvXsylyv&fieldKey=name"
    headers = {
        "Authorization": f"Bearer {api_token}",
    }
    async with httpx.AsyncClient() as client:
        try:
            response = await client.get(url, headers=headers)
            response.raise_for_status()
            return response.json()
        except Exception:
            return None

def convert_to_csv_bytes(data):
    output = io.StringIO()
    writer = csv.writer(output)
    records = data.get('records', [])
    if not records:
        return b""
    headers = records[0]['fields'].keys()
    writer.writerow(headers)
    for record in records:
        writer.writerow([record['fields'].get(h, "") for h in headers])
    return output.getvalue().encode('utf-8')

def convert_to_excel_bytes(data):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    records = data.get('records', [])
    if not records:
        return b""
    headers = list(records[0]['fields'].keys())
    ws.append(headers)
    for record in records:
        ws.append([record['fields'].get(h, "") for h in headers])
    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

async def send_update_to_truetabs(user_id: int):
    tt_config = await get_tt_config(user_id)
    if not tt_config:
        return False
    api_token = decrypt_data(tt_config['api_token'])
    datasheet_id = tt_config['datasheet_id']

    url = f"https://true.tabs.sale/fusion/v1/datasheets/{datasheet_id}/records?viewId=viwyshvXsylyv&fieldKey=name"
    headers = {
        "Authorization": f"Bearer {api_token}",
        "Content-Type": "application/json"
    }
    payload = {
        "records": [],
        "fieldKey": "name"
    }
    async with httpx.AsyncClient() as client:
        try:
            response = await client.patch(url, headers=headers, json=payload)
            response.raise_for_status()
            return True
        except Exception:
            return False
